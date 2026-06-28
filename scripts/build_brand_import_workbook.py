#!/usr/bin/env python3

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook


def normalize_header(value):
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def build_header_index(row):
    header_index = {}
    for cell_index, cell in enumerate(row, start=1):
        normalized = normalize_header(cell.value)
        if normalized and normalized not in header_index:
            header_index[normalized] = cell_index
    return header_index


def detect_header_row(sheet):
    candidates = []
    for row_number in range(1, min(sheet.max_row, 3) + 1):
        header_index = build_header_index(sheet[row_number])
        score = sum(
            1
            for key in ("productkey", "brand", "titleen", "descriptionen", "skucode")
            if key in header_index
        )
        candidates.append((score, row_number, header_index))

    candidates.sort(reverse=True)
    best_score, row_number, header_index = candidates[0]
    if best_score < 2:
        raise ValueError("Unable to detect a valid header row in the source workbook")
    return row_number, header_index


def build_namespaced_value(prefix, value):
    if value in (None, ""):
        return value
    return f"{prefix}-{value}"


def main():
    parser = argparse.ArgumentParser(description="Clone the Sensation import workbook into a target brand workbook.")
    parser.add_argument("--input", required=True, help="Source workbook path")
    parser.add_argument("--output", required=True, help="Output workbook path")
    parser.add_argument("--manifest", required=False, help="Optional JSON manifest path")
    parser.add_argument("--target-brand-key", required=True, help="Target brand key for the Brand column")
    parser.add_argument("--product-limit", type=int, default=12, help="Maximum number of unique product keys to keep")
    parser.add_argument("--identifier-prefix", required=True, help="Prefix used for Product key / SKU / Barcode namespacing")
    parser.add_argument(
        "--keep-source-identifiers",
        action="store_true",
        help="Keep Product key / SKU / Barcode values unchanged",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    manifest_path = Path(args.manifest) if args.manifest else None

    if not input_path.exists():
        raise FileNotFoundError(f"Workbook not found: {input_path}")

    source_workbook = load_workbook(filename=input_path, data_only=False)
    source_sheet = source_workbook.active
    header_row_number, header_index = detect_header_row(source_sheet)

    brand_col = header_index.get("brand")
    product_key_col = header_index.get("productkey")
    sku_code_col = header_index.get("skucode")
    barcode_col = header_index.get("barcode")

    if not brand_col or not product_key_col:
        raise ValueError("The source workbook must include Brand and Product key columns")

    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = source_sheet.title

    header_values = [source_sheet.cell(header_row_number, col).value for col in range(1, source_sheet.max_column + 1)]
    output_sheet.append(header_values)

    seen_product_keys = []
    seen_product_lookup = set()
    rows_written = 0

    for row_number in range(header_row_number + 1, source_sheet.max_row + 1):
        row_values = [source_sheet.cell(row_number, col).value for col in range(1, source_sheet.max_column + 1)]
        raw_product_key = row_values[product_key_col - 1]
        product_key = str(raw_product_key).strip() if raw_product_key not in (None, "") else ""

        if not product_key:
            continue

        if product_key not in seen_product_lookup:
            if args.product_limit > 0 and len(seen_product_keys) >= args.product_limit:
                continue
            seen_product_lookup.add(product_key)
            seen_product_keys.append(product_key)

        row_values[brand_col - 1] = args.target_brand_key

        if not args.keep_source_identifiers:
            row_values[product_key_col - 1] = build_namespaced_value(args.identifier_prefix, product_key)

            if sku_code_col:
                row_values[sku_code_col - 1] = build_namespaced_value(
                    args.identifier_prefix,
                    row_values[sku_code_col - 1],
                )

            if barcode_col and row_values[barcode_col - 1] not in (None, ""):
                row_values[barcode_col - 1] = build_namespaced_value(
                    args.identifier_prefix,
                    row_values[barcode_col - 1],
                )

        output_sheet.append(row_values)
        rows_written += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_workbook.save(output_path)

    manifest = {
        "inputPath": str(input_path),
        "outputPath": str(output_path),
        "targetBrandKey": args.target_brand_key,
        "headerRow": header_row_number,
        "productsWritten": len(seen_product_keys),
        "rowsWritten": rows_written,
        "keptSourceIdentifiers": args.keep_source_identifiers,
        "sampleProductKeys": seen_product_keys[:5],
    }

    if manifest_path:
        manifest_path.parent.mkdir(parents=True, exist_ok=True)
        manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    print(json.dumps(manifest))


if __name__ == "__main__":
    main()
